# ui.py — минималистичный интерфейс (PySide6). Основной файл для запуска.
import sys, calendar
from datetime import date
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QToolBar, QFileDialog, QMessageBox, QLabel
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QBrush, QColor

import db
from logic import generate_schedule

SHIFT_COL_LABEL = "Смена"
DUTY_COL_LABEL = "Деж."

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("График смен/дежурств — прототип")
        db.init_db()
        self.employees = db.load_employees()
        today = date.today()
        self.year = today.year
        self.month = today.month

        self.setup_ui()
        self.load_or_generate()

    def setup_ui(self):
        toolbar = QToolBar()
        self.addToolBar(toolbar)

        self.prev_act = QAction("◀ Месяц", self)
        self.next_act = QAction("Месяц ▶", self)
        self.gen_act  = QAction("Автографик ⚡", self)
        self.save_act = QAction("Сохранить 💾", self)
        self.export_act = QAction("Экспорт Excel ⤓", self)

        toolbar.addAction(self.prev_act)
        self.title_lbl = QLabel("")
        toolbar.addWidget(self.title_lbl)
        toolbar.addAction(self.next_act)
        toolbar.addSeparator()
        toolbar.addAction(self.gen_act)
        toolbar.addAction(self.save_act)
        toolbar.addAction(self.export_act)

        self.prev_act.triggered.connect(self.prev_month)
        self.next_act.triggered.connect(self.next_month)
        self.gen_act.triggered.connect(self.autogenerate)
        self.save_act.triggered.connect(self.save_schedule)
        self.export_act.triggered.connect(self.export_excel)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.cellClicked.connect(self.cell_clicked)

        central = QWidget()
        lay = QVBoxLayout(central)
        lay.addWidget(self.table)
        self.setCentralWidget(central)

    def build_table(self):
        days = [d for d in calendar.Calendar(firstweekday=0).itermonthdates(self.year, self.month) if d.month == self.month]
        self.days = days
        cols = 1 + len(days)*2
        self.table.setColumnCount(cols)
        self.table.setRowCount(len(self.employees))

        headers = ["Сотрудник"]
        for d in days:
            headers += [f"{d.day}\n{SHIFT_COL_LABEL}", f"{d.day}\n{DUTY_COL_LABEL}"]
        self.table.setHorizontalHeaderLabels(headers)

        self.title_lbl.setText(f"{calendar.month_name[self.month]} {self.year}".capitalize())

        for r, e in enumerate(self.employees):
            self.table.setItem(r, 0, QTableWidgetItem(e["name"]))

        # затемнение шапки у выходных
        brush = QBrush(QColor(238,238,238))
        for i, d in enumerate(days):
            if d.weekday() >= 5:
                for sub in (0,1):
                    item = self.table.horizontalHeaderItem(1 + i*2 + sub)
                    if item:
                        item.setBackground(brush)

        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, 180)

    def load_or_generate(self):
        self.build_table()
        loaded = db.load_month_schedule(self.year, self.month)
        if loaded:
            self.schedule = self._merge_loaded(loaded)
        else:
            self.autogenerate()
        self.render_schedule()

    def _merge_loaded(self, loaded):
        result = {}
        for e in self.employees:
            name = e["name"]
            result[name] = {}
            for d in self.days:
                day = d.day
                payload = loaded.get(name, {}).get(day, {"shift":"", "duty":False})
                result[name][day] = {"shift": payload.get("shift","") or "", "duty": bool(payload.get("duty", False))}
        return result

    def render_schedule(self):
        for r, e in enumerate(self.employees):
            name = e["name"]
            for i, d in enumerate(self.days):
                day = d.day
                shift = self.schedule[name][day]["shift"]
                duty = self.schedule[name][day]["duty"]
                self.table.setItem(r, 1 + i*2, QTableWidgetItem(shift))
                self.table.setItem(r, 1 + i*2 + 1, QTableWidgetItem("Д" if duty else ""))

    # -------- actions
    def prev_month(self):
        if self.month == 1:
            self.month = 12; self.year -= 1
        else:
            self.month -= 1
        self.load_or_generate()

    def next_month(self):
        if self.month == 12:
            self.month = 1; self.year += 1
        else:
            self.month += 1
        self.load_or_generate()

    def autogenerate(self):
        self.schedule = generate_schedule(self.employees, self.year, self.month)
        self.render_schedule()
        QMessageBox.information(self, "Готово", "График сгенерирован.")

    def save_schedule(self):
        data = {}
        for r, e in enumerate(self.employees):
            name = e["name"]
            data[name] = {}
            for i, d in enumerate(self.days):
                sh_item = self.table.item(r, 1 + i*2)
                dj_item = self.table.item(r, 1 + i*2 + 1)
                shift = sh_item.text() if sh_item else ""
                duty  = (dj_item.text() == "Д") if dj_item else False
                data[name][d.day] = {"shift": shift, "duty": duty}
        db.save_month_schedule(self.year, self.month, data)
        QMessageBox.information(self, "Сохранено", "Расписание сохранено в базе.")

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Требуется пакет openpyxl: {e}")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Сохранить Excel", f"График_{self.year}_{self.month:02d}.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws = wb.active; ws.title = "График"

        ws.cell(1,1,"Сотрудник"); ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        col = 2
        days = self.days

        fill_weekend = PatternFill("solid", fgColor="EEEEEE")
        fill_shift1  = PatternFill("solid", fgColor="DCEBFF")
        fill_shift2  = PatternFill("solid", fgColor="7FA8F8")
        fill_off     = PatternFill("solid", fgColor="F0F0F0")
        fill_duty    = PatternFill("solid", fgColor="555555")
        font_white   = Font(color="FFFFFF", bold=True)

        for d in days:
            ws.cell(1,col, f"{d.day}"); ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            if d.weekday()>=5:
                ws.cell(1,col).fill = fill_weekend
                ws.cell(2,col).fill = fill_weekend
                ws.cell(2,col+1).fill = fill_weekend
            ws.cell(2,col, "Смена"); ws.cell(2,col+1, "Деж.")
            col += 2

        r = 3
        for e in self.employees:
            name = e["name"]
            ws.cell(r,1,name)
            c = 2
            for d in days:
                sh = self.schedule[name][d.day]["shift"]
                dj = self.schedule[name][d.day]["duty"]
                ws.cell(r,c, sh)
                ws.cell(r,c+1, "Д" if dj else "")
                if sh == "1":
                    ws.cell(r,c).fill = fill_shift1
                elif sh == "2":
                    ws.cell(r,c).fill = fill_shift2
                    ws.cell(r,c).font = font_white
                elif sh == "В":
                    ws.cell(r,c).fill = fill_off
                if dj:
                    ws.cell(r,c+1).fill = fill_duty
                    ws.cell(r,c+1).font = font_white
                c += 2
            r += 1

        wb.save(path)
        QMessageBox.information(self, "Экспорт", f"Файл сохранён: {path}")

    def cell_clicked(self, row, col):
        if col == 0:  # имя
            return
        is_shift_col = (col % 2 == 1)  # 1,3,5... — смена
        day_idx = (col-1)//2
        name = self.employees[row]["name"]
        item = self.table.item(row, col)

        if is_shift_col:
            # цикл: "" -> "1" -> "2" -> "В" -> ""
            cur = (item.text() if item else "") or ""
            nxt = {"": "1", "1": "2", "2":"В", "В":""}.get(cur, "1")
            self.table.setItem(row, col, QTableWidgetItem(nxt))
            # если смена пустая/В — снять дежурство
            self.table.setItem(row, col+1, QTableWidgetItem("" if nxt in ("", "В") else (self.table.item(row, col+1).text() if self.table.item(row, col+1) else "")))
        else:
            # дежурство можно только если соседняя смена 1 или 2
            sh_item = self.table.item(row, col-1)
            sh = sh_item.text() if sh_item else ""
            if sh in ("1","2"):
                cur = (item.text() if item else "") or ""
                nxt = "" if cur == "Д" else "Д"
                # в смене может быть один дежурный → проверим
                if nxt == "Д":
                    for r2 in range(self.table.rowCount()):
                        if r2 == row: continue
                        sh2_item = self.table.item(r2, col-1)
                        dj2_item = self.table.item(r2, col)
                        if (sh2_item and sh2_item.text()==sh) and (dj2_item and dj2_item.text()=="Д"):
                            QMessageBox.warning(self, "Дежурство", "В этой смене уже есть дежурный.")
                            return
                self.table.setItem(row, col, QTableWidgetItem(nxt))
            else:
                QMessageBox.information(self, "Дежурство", "Сначала назначьте смену (1 или 2).")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MainWindow()
    w.resize(1200, 700)
    w.show()
    sys.exit(app.exec())
